import openpyxl
import time


def opera_sheet(wb):

    sheet = wb
    # 操作第一列 第3行到12行的数据
    for i in range(3, 13):
        for j in range(1, 6, 2):
            c = sheet.cell(row=i, column=j)
            # d = sheet.cell(row=i, column=j+1)
            c.value = (i * j + i)
            # d.value = (i * j)

    return sheet


def opera_sheet_headr(wb, colum=1, headers=1):
    # 处理表头
    sheet = wb
    col = colum
    # 操作 第一列 设置为表头
    # 第1列 第2行 list_rD  设置两个曲线的
    for i in range(1, 2):
        for j in range(1, 7, 2):
            c = sheet.cell(row=i, column=j)
            d = sheet.cell(row=i+1, column=j)
            e = sheet.cell(row=i+1, column=j+1)
            c.value = (i * j + i)
            d.value = ('list_rD')
            e.value = ('list_TD')

            # d.value = (i * j)

    return sheet


def opera_sheet_one_colum(wb, row_position=3, column_position=1, body=[]):
    # wb是sheet表
    # row_position =3 column_position = 1 从第1列 第三行 开始
    # body 是传入的数据
    sheet = wb
    rp = row_position
    cp = column_position
    end = len(body)
    # 操作 第一列 设置为表头
    # 第1列 第2行 list_rD  设置两个曲线的 column 列
    p = 0
    for i in range(rp, rp+end):
        for j in range(cp, cp+1):
            c = sheet.cell(row=i, column=j)
            if p < end:
                c.value = body[p]
            p += 1

    return sheet


def demo():
    workbook = openpyxl.Workbook()
    # 获取表格的默认工作表
    sheet = workbook.active
    # 设置工作表的标题 sheet = ws
    # colum 列 row 行，排
    sheet.title = 'demo曲线数据'

    sheet = opera_sheet_headr(sheet)
    sheet = opera_sheet_one_colum(sheet)
    # # 操作第一列 第3行到12行的数据
    # for i in range(3, 13):
    #     for j in range(1, 2):
    #         c = sheet.cell(row=i, column=j)
    #         d = sheet.cell(row=i, column=j+1)
    #         c.value = (i * j)
    #         d.value = (i * j)
    workbook.save('demo数据 .xlsx')


if __name__ == '__main__':
    demo()
