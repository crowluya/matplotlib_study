import json
import openpyxl
import time
from models.td import Td
import numpy as np
'''
需要使用 openpyxl 处理 excel 文件
注意，只支持 xlsx 格式，不支持 xls 格式

需要用 pip 安装 openpyxl 库

openpyxl 的官方文档:
http://openpyxl.readthedocs.io/en/stable/
'''


def log(*args, **kwargs):
    '''
     log为自己写的工具类 打断点 方便调试
    :param args:
    :param kwargs:
    :return:
    '''
    # time.time() 返回 unix time
    # 如何把 unix time 转换为普通人类可以看懂的格式
    format = '%Y/%m/%d %H:%M:%S'
    value = time.localtime(int(time.time()))
    dt = time.strftime(format, value)
    print(dt, *args, **kwargs)


def str_to_str_arr(strs, spo=', '):
    # 接收一个字符串, 返回字符串数组
    # '[57513.48230695434, 69015.9787683452]'
    # 返回['57513.48230695434', '69015.9787683452']
    sp = spo
    list_str = strs
    t = list_str.replace('[', '')
    t = t.replace(']', '')
    arr = t.split(sp)

    return arr


def str_arr_to_float_arr(str_arr):
    # 接收一个字符串数组, 返回浮点数数组
    arr = str_arr
    x = np.array(arr)
    y = x.astype(np.float)

    return y


def str_to_float_arr(strs):
    # 接收一个字符串, 返回浮点数数组
    str_arr = str_to_str_arr(strs)
    f_arr = str_arr_to_float_arr(str_arr)

    return f_arr


def opera_one_td_headr(wb, col=1, title=''):
    # 处理表头
    sheet = wb
    # 表头在第一行
    r = 1

    colu = col
    t = title
    # 操作 第一列 设置为表头
    # 第1列 第2行 list_rD  设置两个曲线的
    for i in range(r, r+1):
        for j in range(colu, colu+1):
            c = sheet.cell(row=i, column=j)
            d = sheet.cell(row=i+1, column=j)
            e = sheet.cell(row=i+1, column=j+1)
            c.value = (t)
            d.value = ('list_rD')
            e.value = ('list_TD')

            # d.value = (i * j)

    return sheet


def opera_one_td_colum(wb, column_position=1, body=[]):
    # wb是sheet表
    # row_position =3 column_position = 1 从第1列 第三行 开始
    # body 是传入的数据
    sheet = wb
    row_position = 3

    # rp 应该大于三
    rp = row_position
    cp = column_position
    end = len(body)

    # 操作 第一列 设置为表头
    # 第1列 第2行 list_rD  设置两个曲线的 column 列
    p = 0
    for i in range(rp, rp+end):
        for j in range(cp, cp+1):
            c = sheet.cell(row=i, column=j)
            if p < end:
                c.value = body[p]
            p += 1

    return sheet


def write_one_td(wb, titl, rd=[], td=[], col=1):
    # 传入sheet
    # 写入rd曲线 写入td曲线的数据
    sheet = wb
    c = col
    t = titl

    sheet = opera_one_td_headr(wb=sheet, col=c, title=t)
    sheet = opera_one_td_colum(wb=sheet, column_position=c, body=rd)
    sheet = opera_one_td_colum(wb=sheet, column_position=c+1, body=td)

    return sheet


def write_xlsx():

    # 1. 加载json文件
    # 2. 将json文件写入到.xlsx中

    # 创建一个新的工作表簿
    all_tds = Td.all()
    # log('all_td ({})'.format(all_tds))
    workbook = openpyxl.Workbook()
    # 获取表格的默认工作表
    sheet = workbook.active
    # 设置工作表的凑题
    sheet.title = '三条tD曲线数据'

    # 写入表头与内容
    column = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    j = 0
    for i in range(len(all_tds)):
        arr_rd = []
        arr_td = []
        td = all_tds[i]
        td = td.json()

        tp = column[j] + str(1)
        rdp = column[j] + str(2)
        tdp = column[j+1] + str(2)

        title = td.get('td', '')
        title = 'td =' + title
        list_rD = td.get('list_rD', '')
        list_TD = td.get('list_TD', '')
        arr_rd = str_to_float_arr(list_rD)
        arr_td = str_to_float_arr(list_TD)

        # 将数据写入sheet
        write_one_td(wb=sheet, rd=arr_rd, td=arr_td, col=j+1, titl=title)
        j += 2
        log('arr_rd({}) arr_td({})'.format(arr_rd, arr_td), i)

    # 保存到 .xlsx 文件中
    workbook.save('({})条曲线数据 .xlsx'.format(len(all_tds)))


def delete_all():
    # 删除所有的td信息
    # 1.返回所有的td类
    # 2.根据td类的id 删除td类(id是不重复的)

    tds = Td.all()

    rs = []
    for t in tds:
        d = Td.delete(t.id)
        rs.append(d)
    log("删除了({})个td, 所有的td= ({})".format(len(tds), rs))
    return "删除了({})个td, 所有的td= ({})".format(len(tds), rs)
    pass


def load_xlsx():
    # 接收一个.xlsx文件
    # 以json格式存储在data/td.txt文件中
    # 本文件可以以字典的形式存储
    # td
    # list_rD=:
    # list_TD=:
    filename = '三条tD曲线数据.xlsx'
    workbook = openpyxl.load_workbook(filename)
    # 所有的工作表单名字
    log(workbook.sheetnames)

    # 拿到某个 工作表 内容的标准做法，我们这个 excel 文件只有一个三条曲线数据工作表
    sheet = workbook['三条曲线数据']
    # sheet['C1'].value     # 返回 C 列第 1 行格子里的值
    # sheet['1']            # 第一行的所有格子(返回一个 tuple)
    # sheet.max_column      # 最大列数
    # sheet.max_row         # 最大行数

    #  获取不同A 列 td的值

    # 用 .value 获取格子里的值
    # log('sheet A1 value', sheet['C1'].value)
    # log('max row', sheet.max_row)
    # 获取一行
    # log('row', sheet['2'])
    # # 遍历第二行所有格子
    # for c in sheet['2']:
    #     log('格子({})'.format(c.value))
    # # 遍历获取 A 列所有的值
    # log('start time')

    le = len(sheet['A'])
    i = 0
    td_cell = []
    for a in sheet['A']:
        # td_cell = []
        v = a.value
        td_cell.append(v)
        i = i + 1
        # log('a', a.value, 'i', i)

    # log('end time cell ({})'.format(td_cell))
    # 每隔六行 新建一个 td类
    dis = 6
    rs = []
    for i in range(0, le, 6):
        cell = []
        dic = {}
        title = td_cell[i].split('=')[1]
        dic['td'] = title
        dic['list_rD'] = td_cell[i + 2]
        dic['list_TD'] = td_cell[i + 4]
        # log('dic', dic, i)
        t = Td.new(dic)
        log('dic', dic, i, t)


'''
 以test开头的是测试函数
 log为自己写的工具类 打断点 方便调试
'''


def test_write():
    write_xlsx()


def test_read():
    load_xlsx()

    # for item in t:
    #     read(item)
    # pass


def test_delete_all():
    delete_all()
    pass


def test():
    # test_read()
    # test_write()
    # test_delete_all()
    pass


def main():
    # read()
    pass


if __name__ == '__main__':
    # main()
    test()
